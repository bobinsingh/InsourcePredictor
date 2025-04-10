from fastapi import APIRouter, HTTPException, UploadFile, File
from typing import List, Dict, Any, Optional
from pydantic import BaseModel
import io
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
import os
from datetime import datetime

from logic.decision_rules import determine_outcome

router = APIRouter()

class DecisionInput(BaseModel):
    business_case: str
    core: str
    legal_requirement: Optional[str] = ""
    risks: Optional[str] = ""
    risk_tolerance: Optional[str] = ""
    frequency: str
    specialised_skill: str
    similarity_with_current_scopes: str
    skill_capacity: str
    duration: str
    affordability: str
    strategic_fit: Optional[str] = ""
    activity_name: str = "Unnamed Activity"
    activity_type: str = ""

class DecisionRequest(BaseModel):
    inputs: List[DecisionInput]

class DecisionOutput(BaseModel):
    activity_name: str
    activity_type: str
    business_case: str
    core: str
    legal_requirement: str
    risks: str
    risk_tolerance: str
    frequency: str
    specialised_skill: str
    similarity_with_current_scopes: str
    skill_capacity: str
    duration: str
    affordability: str
    strategic_fit: str
    outcome: str
    timestamp: Optional[str] = None

class DecisionResponse(BaseModel):
    results: List[DecisionOutput]

@router.post("/determine", response_model=DecisionResponse)
async def determine_outcomes(request: DecisionRequest):
    """
    Determine the outcomes for the provided inputs based on decision rules
    """
    results = []
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    for input_data in request.inputs:
        # Convert input to dict
        input_dict = input_data.dict()
        
        # Determine the outcome
        outcome = determine_outcome(input_dict)
        
        # Create output with all input fields plus outcome
        output = DecisionOutput(
            activity_name=input_dict.get("activity_name", ""),
            activity_type=input_dict.get("activity_type", ""),
            business_case=input_dict.get("business_case", ""),
            core=input_dict.get("core", ""),
            legal_requirement=input_dict.get("legal_requirement", ""),
            risks=input_dict.get("risks", ""),
            risk_tolerance=input_dict.get("risk_tolerance", ""),
            frequency=input_dict.get("frequency", ""),
            specialised_skill=input_dict.get("specialised_skill", ""),
            similarity_with_current_scopes=input_dict.get("similarity_with_current_scopes", ""),
            skill_capacity=input_dict.get("skill_capacity", ""),
            duration=input_dict.get("duration", ""),
            affordability=input_dict.get("affordability", ""),
            strategic_fit=input_dict.get("strategic_fit", ""),
            outcome=outcome,
            timestamp=timestamp
        )
        
        results.append(output)
    
    return DecisionResponse(results=results)

@router.post("/export-excel")
async def export_to_excel(request: DecisionRequest):
    """
    Generate Excel file with inputs and outcomes using openpyxl
    Preserves history when an existing file is uploaded
    """
    # Process inputs to determine outcomes
    results = []
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    for input_data in request.inputs:
        input_dict = input_data.dict()
        outcome = determine_outcome(input_dict)
        
        # Add outcome to the input data
        input_dict["outcome"] = outcome
        input_dict["timestamp"] = timestamp
        results.append(input_dict)
    
    # Create a workbook using openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Sourcing Decisions"
    
    # Define column headers with friendly names
    headers_dict = {
        "activity_name": "Activity Name", 
        "activity_type": "Activity Type",
        "business_case": "Business case",
        "core": "Core ", 
        "legal_requirement": "Legal requirement", 
        "risks": "Risks", 
        "risk_tolerance": "Risk tolerance ", 
        "frequency": "Frequency", 
        "specialised_skill": "Specialised Skill", 
        "similarity_with_current_scopes": "Similarity with current scopes", 
        "skill_capacity": "Skill capacity", 
        "duration": "Duration ", 
        "affordability": "Affordability & Transferable Skill", 
        "strategic_fit": "Strategic fit and Business case", 
        "outcome": "Outcome",
        "timestamp": "Timestamp"
    }
    
    # Define column order
    headers = [
        "activity_name", "activity_type", "business_case", "core", "legal_requirement", 
        "risks", "risk_tolerance", "frequency", "specialised_skill", 
        "similarity_with_current_scopes", "skill_capacity", "duration", 
        "affordability", "strategic_fit", "outcome", "timestamp"
    ]
    
    # Add headers with styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=headers_dict[header])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    for row_idx, result in enumerate(results, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=result.get(header, ""))
            
            # Color the outcome cell based on the value
            if header == "outcome":
                if result[header] == "Eliminate":
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                elif result[header] == "Current Outsource":
                    cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                elif result[header] == "New Outsource":
                    cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                elif result[header] == "Insource or create in-house capacity":
                    cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                elif result[header] == "Requires Further Analysis":
                    cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    # Add explanatory notes as a separate sheet
    notes_sheet = wb.create_sheet(title="Field Descriptions")
    
    # Add field descriptions
    field_descriptions = {
        "Business case": "Does this activity have a business case?",
        "Core ": "Core activities are essential to your organization's primary mission and competitive advantage",
        "Legal requirement": "Activities required by law, regulation, or contract that cannot be eliminated",
        "Risks": "Consider reputational, operational, financial, or compliance risks associated with this activity",
        "Risk tolerance ": "Inside or outside your organization's risk tolerance boundaries",
        "Frequency": "How often the activity is performed (High/Low)",
        "Specialised Skill": "Whether the activity requires specialized expertise or capabilities",
        "Similarity with current scopes": "How similar this activity is to your organization's existing operations and capabilities",
        "Skill capacity": "Whether your organization already has the necessary skills and resources to perform this activity",
        "Duration ": "The expected timeframe for this activity (Short = temporary or project-based, Long = ongoing or permanent)",
        "Affordability & Transferable Skill": "Whether your organization can afford to develop or maintain this capability internally",
        "Strategic fit and Business case": "How well this activity aligns with your organization's long-term strategic objectives",
        "Outcome": "The recommended sourcing strategy based on all factors",
        "Timestamp": "Date and time when the decision was made"
    }
    
    # Add header
    notes_sheet["A1"] = "Field"
    notes_sheet["B1"] = "Description"
    notes_sheet["A1"].font = header_font
    notes_sheet["B1"].font = header_font
    notes_sheet["A1"].fill = header_fill
    notes_sheet["B1"].fill = header_fill
    
    # Add descriptions
    row = 2
    for field, description in field_descriptions.items():
        notes_sheet[f"A{row}"] = field
        notes_sheet[f"B{row}"] = description
        row += 1
    
    # Adjust column width
    notes_sheet.column_dimensions["A"].width = 30
    notes_sheet.column_dimensions["B"].width = 100
    
    # Add a history sheet to record all decisions
    history_sheet = wb.create_sheet(title="Decision History")
    
    # Add headers to history sheet
    for col_idx, header in enumerate(headers, 1):
        cell = history_sheet.cell(row=1, column=col_idx, value=headers_dict[header])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add current results to history
    history_row = 2
    for result in results:
        for col_idx, header in enumerate(headers, 1):
            history_sheet.cell(row=history_row, column=col_idx, value=result.get(header, ""))
        history_row += 1
    
    # Auto-adjust column width for history sheet
    for col in history_sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
        adjusted_width = (max_length + 2) * 1.2
        history_sheet.column_dimensions[column].width = adjusted_width
    
    # Save to a BytesIO object
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return the Excel file as a response
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sourcing_decisions.xlsx"}
    )

@router.post("/update-excel")
async def update_excel(request: DecisionRequest, existing_file: UploadFile = File(...)):
    """
    Update an existing Excel file with new decisions without overwriting history
    """
    try:
        # Read the existing Excel file
        contents = await existing_file.read()
        existing_wb = openpyxl.load_workbook(io.BytesIO(contents))
        
        # Process inputs to determine outcomes
        results = []
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        for input_data in request.inputs:
            input_dict = input_data.dict()
            outcome = determine_outcome(input_dict)
            
            # Add outcome and timestamp to the input data
            input_dict["outcome"] = outcome
            input_dict["timestamp"] = timestamp
            results.append(input_dict)
        
        # Get the history sheet or create if it doesn't exist
        if "Decision History" in existing_wb.sheetnames:
            history_sheet = existing_wb["Decision History"]
        else:
            history_sheet = existing_wb.create_sheet(title="Decision History")
            
            # Define headers
            headers = [
                "activity_name", "activity_type", "business_case", "core", "legal_requirement", 
                "risks", "risk_tolerance", "frequency", "specialised_skill", 
                "similarity_with_current_scopes", "skill_capacity", "duration", 
                "affordability", "strategic_fit", "outcome", "timestamp"
            ]
            
            headers_dict = {
                "activity_name": "Activity Name", 
                "activity_type": "Activity Type",
                "business_case": "Business case",
                "core": "Core ", 
                "legal_requirement": "Legal requirement", 
                "risks": "Risks", 
                "risk_tolerance": "Risk tolerance ", 
                "frequency": "Frequency", 
                "specialised_skill": "Specialised Skill", 
                "similarity_with_current_scopes": "Similarity with current scopes", 
                "skill_capacity": "Skill capacity", 
                "duration": "Duration ", 
                "affordability": "Affordability & Transferable Skill", 
                "strategic_fit": "Strategic fit and Business case", 
                "outcome": "Outcome",
                "timestamp": "Timestamp"
            }
            
            # Add headers to history sheet
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            
            for col_idx, header in enumerate(headers, 1):
                cell = history_sheet.cell(row=1, column=col_idx, value=headers_dict[header])
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Find the last row in the history sheet
        last_row = history_sheet.max_row
        if last_row < 1:  # In case there's an empty sheet
            last_row = 1
        
        # Add new results to history
        headers = [
            "activity_name", "activity_type", "business_case", "core", "legal_requirement", 
            "risks", "risk_tolerance", "frequency", "specialised_skill", 
            "similarity_with_current_scopes", "skill_capacity", "duration", 
            "affordability", "strategic_fit", "outcome", "timestamp"
        ]
        
        for idx, result in enumerate(results):
            for col_idx, header in enumerate(headers, 1):
                history_sheet.cell(row=last_row + 1 + idx, column=col_idx, value=result.get(header, ""))
        
        # Update the main sheet with the new results
        main_sheet = existing_wb["Sourcing Decisions"]
        
        # Clear the current data (except headers)
        for row in range(2, main_sheet.max_row + 1):
            for col in range(1, main_sheet.max_column + 1):
                main_sheet.cell(row=row, column=col).value = None
                main_sheet.cell(row=row, column=col).fill = PatternFill(fill_type=None)
        
        # Add the new data
        for row_idx, result in enumerate(results, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = main_sheet.cell(row=row_idx, column=col_idx, value=result.get(header, ""))
                
                # Color the outcome cell based on the value
                if header == "outcome":
                    if result[header] == "Eliminate":
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    elif result[header] == "Current Outsource":
                        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                    elif result[header] == "New Outsource":
                        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                    elif result[header] == "Insource or create in-house capacity":
                        cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                    elif result[header] == "Requires Further Analysis":
                        cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Save the updated workbook
        output = io.BytesIO()
        existing_wb.save(output)
        output.seek(0)
        
        # Return the updated Excel file as a response
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={existing_file.filename}"}
        )
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating Excel file: {str(e)}")